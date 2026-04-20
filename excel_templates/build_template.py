"""
build_template.py — generate a KAT pivot-starter workbook.

Produces `kat_pivot_starter.xlsx` with four tabs:
  Data         — the raw transaction data, already as an Excel Table
  LineSummary  — per-line loss / expense / combined ratios, formula-driven
  MonthlyTrend — monthly loss ratio time series
  RegionMatrix — region × product_line loss ratio matrix
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "..", "data", "kat_insurance_sales.csv")
OUT = os.path.join(BASE, "kat_pivot_starter.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1f4e78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def widen(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    df = pd.read_csv(DATA, parse_dates=["transaction_date"])
    df["year"] = df["transaction_date"].dt.year
    df["month"] = df["transaction_date"].dt.month

    wb = Workbook()

    # --- Data sheet ---
    ws = wb.active
    ws.title = "Data"
    # Write a manageable sample (full dataset is ~67k rows; Excel tolerates it fine
    # but we keep 20k for template speed)
    sample = df.sample(n=min(20000, len(df)), random_state=3).reset_index(drop=True)
    cols = [c for c in sample.columns]
    ws.append(cols)
    for row in dataframe_to_rows(sample, index=False, header=False):
        ws.append(row)
    style_header(ws, 1, len(cols))
    tab = Table(displayName="DataTbl", ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)
    widen(ws, {c: 14 for c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"})
    ws.freeze_panes = "A2"

    # --- LineSummary ---
    ws = wb.create_sheet("LineSummary")
    ws.append(["Product Line", "Net Premium", "Claims", "Commission", "Admin",
               "Loss Ratio", "Expense Ratio", "Combined Ratio", "UW Margin"])
    style_header(ws, 1, 9)
    lines = sorted(df["product_line"].unique())
    for i, line in enumerate(lines, start=2):
        sub = df[df["product_line"] == line]
        np_ = sub["net_premium"].sum()
        cl = sub["claim_amount"].sum()
        co = sub["commission"].sum()
        ad = sub["admin_cost"].sum()
        ws.append([line, np_, cl, co, ad, None, None, None, None])
        # Formulas reference the row cells
        ws.cell(row=i, column=6).value = f"=C{i}/B{i}"
        ws.cell(row=i, column=7).value = f"=(D{i}+E{i})/B{i}"
        ws.cell(row=i, column=8).value = f"=F{i}+G{i}"
        ws.cell(row=i, column=9).value = f"=1-H{i}"
        for c in [6, 7, 8, 9]:
            ws.cell(row=i, column=c).number_format = "0.0%"
        for c in [2, 3, 4, 5]:
            ws.cell(row=i, column=c).number_format = '"$"#,##0'
    # Conditional formatting on combined ratio — green below 100%, red above
    ws.conditional_formatting.add(
        f"H2:H{1+len(lines)}",
        ColorScaleRule(
            start_type="num", start_value=0.7, start_color="2ECC71",
            mid_type="num", mid_value=1.0, mid_color="F1C40F",
            end_type="num", end_value=1.3, end_color="E74C3C",
        ),
    )
    widen(ws, {"A": 16, "B": 16, "C": 16, "D": 16, "E": 14, "F": 13, "G": 15, "H": 16, "I": 13})

    # --- MonthlyTrend ---
    ws = wb.create_sheet("MonthlyTrend")
    ws.append(["Year", "Month", "Net Premium", "Claims", "Loss Ratio"])
    style_header(ws, 1, 5)
    monthly = (
        df.groupby(["year", "month"])
        .agg(np_=("net_premium", "sum"), cl=("claim_amount", "sum"))
        .reset_index()
    )
    for i, r in enumerate(monthly.itertuples(index=False), start=2):
        ws.append([r.year, r.month, r.np_, r.cl, None])
        ws.cell(row=i, column=3).number_format = '"$"#,##0'
        ws.cell(row=i, column=4).number_format = '"$"#,##0'
        ws.cell(row=i, column=5).value = f"=D{i}/C{i}"
        ws.cell(row=i, column=5).number_format = "0.0%"
    ws.conditional_formatting.add(
        f"E2:E{1+len(monthly)}",
        ColorScaleRule(
            start_type="min", start_color="2ECC71",
            mid_type="percentile", mid_value=50, mid_color="F1C40F",
            end_type="max", end_color="E74C3C",
        ),
    )
    widen(ws, {"A": 8, "B": 8, "C": 16, "D": 16, "E": 13})

    # --- RegionMatrix ---
    ws = wb.create_sheet("RegionMatrix")
    region_matrix = (
        df.groupby(["region", "product_line"])
        .apply(lambda g: g["claim_amount"].sum() / g["net_premium"].sum())
        .unstack("product_line")
    )
    cols = ["Region"] + list(region_matrix.columns)
    ws.append(cols)
    style_header(ws, 1, len(cols))
    for i, (region, row) in enumerate(region_matrix.iterrows(), start=2):
        ws.append([region] + list(row.values))
        for c in range(2, len(cols) + 1):
            ws.cell(row=i, column=c).number_format = "0.0%"
    end_col_letter = chr(ord("A") + len(region_matrix.columns))
    ws.conditional_formatting.add(
        f"B2:{end_col_letter}{1+len(region_matrix)}",
        ColorScaleRule(
            start_type="min", start_color="2ECC71",
            mid_type="percentile", mid_value=50, mid_color="F1C40F",
            end_type="max", end_color="E74C3C",
        ),
    )
    widen(ws, {"A": 18, **{chr(ord("B") + i): 14 for i in range(len(region_matrix.columns))}})

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
